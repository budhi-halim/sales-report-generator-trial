import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import io
from datetime import datetime
from pyscript import ffi, window, document
import re
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Section: Configuration Class
# Holds all configurable settings for columns, formats, and processing rules
class ColumnConfig:
    def __init__(self):
        self.main_column = 'B'  # Main column to check for data presence
        self.extract_columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']  # Columns to extract
        self.validation_column = 'K'  # Column to validate for blanks in rows 4-6
        self.key_map = {  # Mapping from column letters to data keys
            'B': 'area',
            'C': 'date',
            'D': 'invoice_no',
            'E': 'customer_name',
            'F': 'product_type',
            'G': 'product_name',
            'H': 'quantity',
            'I': 'unit_price'
        }
        self.output_headers = [  # Headers for the output sales sheet
            'Area',
            'Tanggal',
            'Periode',
            'No. Faktur',
            'Customer',
            'Jenis',
            'Produk',
            'Jumlah',
            'Harga Satuan',
            'Total (IDR)',
            'Kurs',
            'Total (USD)'
        ]
        self.preserve_upper_customer = {  # Words to keep uppercase for customers
            'ABC',
            'BAL',
            'TSG'
        }
        self.preserve_upper_product = {  # Words to keep uppercase for products
            'GMS',
            'HVP',
            'WCI',
            'BBQ',
            'KAN'
        }
        self.area_replacements = {  # Area abbreviations to full names
            'Bdg': 'Bandung',
            'Bgr': 'Bogor',
            'Bks': 'Bekasi',
            'Jkt': 'Jakarta',
            'Lpg': 'Lampung',
            'Mdn': 'Medan',
            'Tgr': 'Tangerang'
        }
        self.idr_format_2 = r'_-"Rp"* #,##0.00_ ;_-"Rp"* -#,##0.00_ ;_-"Rp"* "-"??_ ;_-@_-'
        self.idr_format_0 = r'_-"Rp"* #,##0_ ;_-"Rp"* -#,##0_ ;_-"Rp"* "-"??_ ;_-@_-'
        self.usd_format_2 = r'_-"$"* #,##0.00_ ;_-"$"* -#,##0.00_ ;_-"Rp"* "-"??_ ;_-@_-' 

# Section: Utility Functions
# Helper functions for data processing
def col_to_index(col):
    # Converts column letter to zero-based index
    return ord(col.upper()) - ord('A')

def proper_case(text, preserve_upper=set()):
    if not text:
        return ''
    # Split into alphanum words and non-alphanum separators
    parts = re.findall(r'[a-zA-Z0-9]+|[^a-zA-Z0-9]+', str(text).strip())
    processed_parts = []
    for part in parts:
        if re.match(r'^[a-zA-Z0-9]+$', part):  # word
            # Split into letter and digit subparts
            subparts = re.findall(r'[a-zA-Z]+|[0-9]+', part)
            processed_sub = []
            for sub in subparts:
                if sub.isdigit():
                    processed_sub.append(sub)
                else:  # letters
                    upper_sub = sub.upper()
                    if upper_sub in preserve_upper:
                        processed_sub.append(upper_sub)
                    elif len(sub) <= 2:
                        processed_sub.append(sub)  # preserve case
                    else:
                        processed_sub.append(sub.capitalize())
            processed_parts.append(''.join(processed_sub))
        else:  # separator
            processed_parts.append(part)
    return ''.join(processed_parts)

def process_area(area, replacements):
    # Processes area string with replacements and casing
    if not area:
        return ''
    area = proper_case(area.strip().split()[0])
    return replacements.get(area, area)

# Section: Structure Validation
# Validates input file structure for blanks in validation column
def validate_structure(buffer, filename, config):
    # Checks if specified validation cells are blank
    engine = 'xlrd' if filename.endswith('.xls') else 'openpyxl'
    df = pd.read_excel(buffer, header=None, engine=engine)
    validation_idx = col_to_index(config.validation_column)
    if validation_idx >= df.shape[1]:
        return True  # Column doesn't exist, treat as valid
    cells = [df.iloc[3, validation_idx], df.iloc[4, validation_idx], df.iloc[5, validation_idx]]
    return all(pd.isna(cell) or str(cell).strip() == '' for cell in cells)

# Section: Date Validation
# Validates extracted dates are proper datetime
def validate_dates(data):
    # Checks if all dates in extracted data are valid datetime
    for row in data:
        date_val = row.get('date')
        if date_val and not isinstance(date_val, datetime):
            # Try force convert if str or num
            try:
                if isinstance(date_val, (str, int, float)):
                    pd.to_datetime(date_val)
                else:
                    return False
            except:
                return False
    return True

# Section: Data Extraction
# Extracts relevant data from input file
def extract_data(buffer, filename, config):
    # Reads and extracts data rows based on config, handling date conversion
    engine = 'xlrd' if filename.endswith('.xls') else 'openpyxl'
    df = pd.read_excel(buffer, header=None, engine=engine, skiprows=3)
    main_idx = col_to_index(config.main_column)
    extract_idxs = [col_to_index(col) for col in config.extract_columns]
    data = []
    for _, row in df.iterrows():
        if pd.isna(row[main_idx]) or str(row[main_idx]).strip() == '':
            continue
        row_data = {}
        for i, col in enumerate(config.extract_columns):
            val = row[extract_idxs[i]]
            if pd.isna(val):
                val = ''
            elif col == 'H':
                try:
                    val = float(val)
                except:
                    val = ''
            elif col == 'I':
                try:
                    val = float(val)
                except:
                    val = ''
            elif col == 'C':
                if isinstance(val, (str, int, float)):
                    try:
                        val = pd.to_datetime(val)
                    except:
                        val = val  # Keep original if fail, validate later
            elif col == 'D':
                val = str(val).rstrip('.0') if isinstance(val, float) else str(val)  # Force str, remove .0
            row_data[config.key_map[col]] = val
        data.append(row_data)
    return data

# Section: Data Processing
# Normalizes and processes extracted data
def process_data(combined_data, config):
    # Applies normalization and collects unique periods
    periods = set()
    for row in combined_data:
        row['area'] = process_area(row['area'], config.area_replacements)
        row['customer_name'] = proper_case(row['customer_name'], config.preserve_upper_customer)
        row['product_name'] = proper_case(row['product_name'], config.preserve_upper_product)
        date_value = row.get('date')
        if date_value and isinstance(date_value, datetime):
            periods.add(date_value.strftime('%Y-%m'))
        row['invoice_no'] = str(row['invoice_no']) if row.get('invoice_no') else ''
        row['product_type'] = str(row['product_type']) if row.get('product_type') else ''
        row['quantity'] = row['quantity'] if row.get('quantity') else ''
        row['unit_price'] = row['unit_price'] if row.get('unit_price') else ''
    combined_data = sorted(combined_data, key=lambda x: x.get('date') or datetime.min)
    return sorted(list(periods)), combined_data

# Section: Blank Check
# Checks for blank required fields
def check_blanks(combined_data):
    # Identifies blank cells in required fields
    warnings = []
    key_to_col = {
        'area': 'A',
        'date': 'B',
        'invoice_no': 'D',
        'customer_name': 'E',
        'product_type': 'F',
        'product_name': 'G',
        'quantity': 'H',
        'unit_price': 'I'
    }
    required_keys = list(key_to_col.keys())
    for i, row in enumerate(combined_data, 1):
        for key in required_keys:
            if not row.get(key):
                warnings.append(f"{key_to_col[key]}{i+1}")
    return warnings

# Section: Table Generation
# General function to generate Excel table from 2D data
def generate_table(sheet, headers, data_2d, table_name, col_formats=None, row_formulas=None, start_row=1, start_col=1):
    # Writes headers and data to sheet, applies formats/formulas, creates table
    for rel_c, header in enumerate(headers, 1):
        c = start_col + rel_c - 1
        sheet.cell(start_row, c).value = header
    for r_idx, row_data in enumerate(data_2d, 0):
        r = start_row + r_idx + 1
        for rel_c, val in enumerate(row_data, 1):
            c = start_col + rel_c - 1
            sheet.cell(r, c).value = val
            if col_formats and rel_c in col_formats:
                sheet.cell(r, c).number_format = col_formats[rel_c]
        if row_formulas and row_formulas.get(r_idx):
            for rel_col, formula in row_formulas[r_idx].items():
                c = start_col + rel_col - 1
                sheet.cell(r, c).value = formula
                if col_formats and rel_col in col_formats:
                    sheet.cell(r, c).number_format = col_formats[rel_col]
    last_col = start_col + len(headers) - 1
    last_row = start_row + len(data_2d)
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(last_col)}{last_row}"
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    sheet.add_table(tab)

# Section: Group Sheet Generation
# General function to generate group-based sheets (Customer/Area - IDR/USD/Qty)
def generate_group_sheet(wb, group_key, unit, combined_data, period_columns, config):
    sort_unit = 'IDR' if unit == 'USD' else unit
    group_to_total = defaultdict(float)
    for row in combined_data:
        q = row['quantity']
        if not isinstance(q, (float, int)):
            continue
        if sort_unit == 'Qty':
            group_to_total[row[group_key]] += q
        else:
            u = row['unit_price']
            if isinstance(u, (float, int)):
                group_to_total[row[group_key]] += q * u
    groups = sorted(group_to_total.keys(), key=lambda g: -group_to_total[g])
    if group_key == 'customer_name':
        sheet_name = f'Customer - {unit}'
        header_label = 'Customer'
        table_name = f'Customer_{unit}'
        group_col = 'Customer'
    elif group_key == 'area':
        sheet_name = f'Area - {unit}'
        header_label = 'Area'
        table_name = f'Area_{unit}'
        group_col = 'Area'
    elif group_key == 'product_name':
        sheet_name = f'Product - {unit}'
        header_label = 'Product'
        table_name = f'Product_{unit}'
        group_col = 'Produk'
    sheet = wb.create_sheet(sheet_name)
    headers = [header_label] + [pc['label'] for pc in period_columns]
    data_2d = [[group] + [''] * len(period_columns) for group in groups]
    col_formats = {}
    col_formats[1] = '@'
    if unit == 'IDR':
        for col in range(2, len(headers)+1):
            col_formats[col] = config.idr_format_2
    elif unit == 'USD':
        for col in range(2, len(headers)+1):
            col_formats[col] = config.usd_format_2
    elif unit == 'Qty':
        for col in range(2, len(headers)+1):
            col_formats[col] = '#,##0'
    row_formulas = {}
    for r_idx in range(len(groups)):
        row_formulas[r_idx] = {}
        r = r_idx + 2
        for pc_idx, pc in enumerate(period_columns):
            c = pc_idx + 2
            if pc['type'] == 'month':
                per = pc['periods'][0]
                y, m = map(int, per.split('-'))
                date_str = f'DATE({y},{m},1)'
                if unit == 'IDR':
                    sum_col = 'Total (IDR)'
                elif unit == 'USD':
                    sum_col = 'Total (USD)'
                elif unit == 'Qty':
                    sum_col = 'Jumlah'
                formula = f'=SUMIFS(Sales_Data[{sum_col}], Sales_Data[{group_col}], $A{r}, Sales_Data[Periode], {date_str})'
            else:
                sum_cols = pc['sum_month_cols']
                sum_str = ','.join(f'{sc}{r}' for sc in sum_cols)
                formula = f'=SUM({sum_str})'
            row_formulas[r_idx][c] = formula
    generate_table(sheet, headers, data_2d, table_name, col_formats, row_formulas)
    end_row = 1 + len(data_2d)
    total_row = end_row + 2
    sheet.cell(total_row, 1).value = 'Total'
    sheet.cell(total_row, 1).font = Font(bold=True)
    for rel_c in range(2, len(headers)+1):
        c = 1 + rel_c - 1
        sum_formula = f'=SUM({get_column_letter(c)}2:{get_column_letter(c)}{end_row})'
        sheet.cell(total_row, c).value = sum_formula
        if col_formats and rel_c in col_formats:
            sheet.cell(total_row, c).number_format = col_formats[rel_c]
        sheet.cell(total_row, c).font = Font(bold=True)

def compute_yearly_sort_totals(combined_data, group_key, unit):
    sort_unit = 'IDR' if unit == 'USD' else unit
    yearly_totals = defaultdict(lambda: defaultdict(float))  # year -> group -> total
    for row in combined_data:
        date = row.get('date')
        if not isinstance(date, datetime):
            continue
        year = str(date.year)
        group = row[group_key]
        q = row['quantity']
        if not isinstance(q, (int, float)):
            continue
        if sort_unit == 'Qty':
            yearly_totals[year][group] += q
        else:  # IDR or USD (using IDR)
            u = row['unit_price']
            if isinstance(u, (int, float)):
                yearly_totals[year][group] += q * u
    return yearly_totals

def generate_yearly_group_sheet(wb, group_key, unit, combined_data, period_columns, config, year_to_col_letter):
    if group_key == 'customer_name':
        sheet_name = f'Customer - Yearly - {unit}'
        header_label = 'Customer'
        table_base = f'Customer_{unit}'
        orig_sheet = f'Customer - {unit}'
    elif group_key == 'area':
        sheet_name = f'Area - Yearly - {unit}'
        header_label = 'Area'
        table_base = f'Area_{unit}'
        orig_sheet = f'Area - {unit}'
    elif group_key == 'product_name':
        sheet_name = f'Product - Yearly - {unit}'
        header_label = 'Product'
        table_base = f'Product_{unit}'
        orig_sheet = f'Product - {unit}'
    sheet = wb.create_sheet(sheet_name)
    yearly_totals = compute_yearly_sort_totals(combined_data, group_key, unit)
    all_groups = sorted(set(row[group_key] for row in combined_data if row[group_key]))
    col_formats = {1: '@'}
    if unit == 'IDR':
        col_formats[2] = config.idr_format_2
    elif unit == 'USD':
        col_formats[2] = config.usd_format_2
    elif unit == 'Qty':
        col_formats[2] = '#,##0'
    current_start_col = 1
    for year in sorted(yearly_totals.keys()):
        year_totals = yearly_totals[year]
        groups = sorted(all_groups, key=lambda g: -year_totals.get(g, 0))
        headers = [header_label, f'Total {year}']
        data_2d = [[g, ''] for g in groups]
        table_name = f'{table_base}_{year}'
        row_formulas = {}
        for r_idx in range(len(groups)):
            row_formulas[r_idx] = {}
            r = r_idx + 2
            formula = f"=INDEX('{orig_sheet}'!{year_to_col_letter[year]}:{year_to_col_letter[year]}, MATCH({get_column_letter(current_start_col)}{r}, '{orig_sheet}'!A:A, 0))"
            row_formulas[r_idx][2] = formula
        generate_table(sheet, headers, data_2d, table_name, col_formats, row_formulas, start_col=current_start_col)
        end_row = 1 + len(data_2d)
        total_row = end_row + 2
        sheet.cell(total_row, current_start_col).value = 'Total'
        sheet.cell(total_row, current_start_col).font = Font(bold=True)
        for rel_c in range(2, len(headers)+1):
            c = current_start_col + rel_c - 1
            sum_formula = f'=SUM({get_column_letter(c)}2:{get_column_letter(c)}{end_row})'
            sheet.cell(total_row, c).value = sum_formula
            if col_formats and rel_c in col_formats:
                sheet.cell(total_row, c).number_format = col_formats[rel_c]
            sheet.cell(total_row, c).font = Font(bold=True)
        current_start_col += 3

# Section: Main Processing Function
# Orchestrates file processing and output
def process_files(file_datas):
    message_div = document.getElementById('message')
    message_div.innerHTML = ''
    file_datas = file_datas.to_py()
    if not file_datas:
        message = 'No files uploaded.'
        message_div.innerHTML = message
        return {'message': message}
    config = ColumnConfig()
    invalid_files = []  # Structure issues
    invalid_date_files = []  # Date format issues
    combined_data = []
    for f in file_datas:
        name = f['name']
        buffer = io.BytesIO(bytes(f['data']))
        if not validate_structure(buffer, name, config):
            invalid_files.append(name)
            continue
        buffer.seek(0)
        data = extract_data(buffer, name, config)
        if not validate_dates(data):
            invalid_date_files.append(name)
            continue
        combined_data.extend(data)
    msg = ''
    if invalid_files:
        msg += '<p>These files do not follow the required format:</p><ul>' + ''.join(f'<li>{f}</li>' for f in invalid_files) + '</ul>'
    if invalid_date_files:
        if msg:
            msg += ''
        msg += '<p>These files have invalid date formatting:</p><ul>' + ''.join(f'<li>{f}</li>' for f in invalid_date_files) + '</ul>'
    if msg:
        message_div.innerHTML = msg
        return {'message': msg, 'type': 'error'}
    sorted_periods, combined_data = process_data(combined_data, config)
    blank_cells = check_blanks(combined_data)
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = 'Sales Data'
    # Build 2D data for sales
    sales_data_2d = []
    sales_row_formulas = {}
    sales_col_formats = {
        1: '@',
        2: 'dd/mm/yyyy',
        3: 'dd/mm/yyyy',
        4: '@',
        5: '@',
        6: '@',
        7: '@',
        8: '#,##0',
        9: config.idr_format_2,
        10: config.idr_format_2,
        11: config.idr_format_0,
        12: config.usd_format_2
    }
    for i, row in enumerate(combined_data):
        date_value = row.get('date') if isinstance(row.get('date'), datetime) else ''
        sales_data_2d.append([row['area'], date_value, '', row['invoice_no'], row['customer_name'], row['product_type'], row['product_name'], row['quantity'], row['unit_price'], '', '', ''])
        sales_row_formulas[i] = {
            3: f'=DATE(YEAR(B{i+2}), MONTH(B{i+2}), 1)',
            10: f'=PRODUCT(H{i+2},I{i+2})',
            11: f'=VLOOKUP(TEXT(C{i+2}, "YYYY-MM"), \'Exchange Rate\'!A:B, 2, FALSE)',
            12: f'=J{i+2}/K{i+2}'
        }
    generate_table(sheet1, config.output_headers, sales_data_2d, "Sales_Data", sales_col_formats, sales_row_formulas)
    sheet2 = wb.create_sheet('Exchange Rate')
    # Build 2D data for exchange (Python-computed periods)
    exchange_headers = ['Period', 'Rate']
    exchange_data_2d = [[p, None] for p in sorted_periods]
    exchange_col_formats = {1: '@'}
    generate_table(sheet2, exchange_headers, exchange_data_2d, "Exchange_Rate", exchange_col_formats)
    # Build period_columns
    years = sorted(set(p[:4] for p in sorted_periods))
    period_columns = []
    current_col = 2
    all_month_letters = []
    year_month_letters = defaultdict(list)
    for year in years:
        months_present = sorted(int(p[5:]) for p in sorted_periods if p[:4] == year)
        for q in range(1, 5):
            q_start = (q - 1) * 3 + 1
            q_end = q_start + 2
            q_months = [m for m in months_present if q_start <= m <= q_end]
            if not q_months:
                continue
            q_letters = []
            for m in q_months:
                label = datetime(int(year), m, 1).strftime('%b %Y')
                letter = get_column_letter(current_col)
                period_columns.append({
                    'label': label,
                    'type': 'month',
                    'periods': [f'{year}-{m:02d}'],
                    'sum_month_cols': None
                })
                q_letters.append(letter)
                all_month_letters.append(letter)
                year_month_letters[year].append(letter)
                current_col += 1
            q_label = f'Q{q} {year}'
            letter = get_column_letter(current_col)
            period_columns.append({
                'label': q_label,
                'type': 'quarter',
                'periods': [f'{year}-{mm:02d}' for mm in q_months],
                'sum_month_cols': q_letters
            })
            current_col += 1
        if months_present:
            label = f'Total {year}'
            letter = get_column_letter(current_col)
            period_columns.append({
                'label': label,
                'type': 'year',
                'periods': [f'{year}-{m:02d}' for m in months_present],
                'sum_month_cols': year_month_letters[year]
            })
            current_col += 1
    if sorted_periods:
        label = 'Total'
        letter = get_column_letter(current_col)
        period_columns.append({
            'label': label,
            'type': 'grand',
            'periods': sorted_periods,
            'sum_month_cols': all_month_letters
        })
    year_to_col_letter = {}
    for idx, pc in enumerate(period_columns):
        if pc['type'] == 'year':
            year = pc['label'].split()[1]
            col_idx = idx + 2
            year_to_col_letter[year] = get_column_letter(col_idx)
    for group_key in ['customer_name', 'area', 'product_name']:
        for unit in ['IDR', 'USD', 'Qty']:
            generate_group_sheet(wb, group_key, unit, combined_data, period_columns, config)
            generate_yearly_group_sheet(wb, group_key, unit, combined_data, period_columns, config, year_to_col_letter)
    # Summary
    period_labels = []
    current_row_idx = 0
    all_month_row_nums = []
    year_month_row_nums = defaultdict(list)
    row_to_sum_rows = {}
    for year in years:
        months_present = sorted(int(p[5:]) for p in sorted_periods if p[:4] == year)
        has_year = False
        for q in range(1, 5):
            q_start = (q - 1) * 3 + 1
            q_end = q_start + 2
            q_months = [m for m in months_present if q_start <= m <= q_end]
            if not q_months:
                continue
            has_year = True
            q_month_rows = []
            for m in q_months:
                label = datetime(int(year), m, 1).strftime('%b %Y')
                period_labels.append(label)
                row_num = current_row_idx + 2
                q_month_rows.append(row_num)
                all_month_row_nums.append(row_num)
                year_month_row_nums[year].append(row_num)
                current_row_idx += 1
            q_label = f'Q{q} {year}'
            period_labels.append(q_label)
            q_row_num = current_row_idx + 2
            row_to_sum_rows[q_row_num] = q_month_rows
            current_row_idx += 1
        if has_year:
            y_label = f'Total {year}'
            period_labels.append(y_label)
            y_row_num = current_row_idx + 2
            row_to_sum_rows[y_row_num] = year_month_row_nums[year]
            current_row_idx += 1
    if sorted_periods:
        grand_label = 'Total'
        period_labels.append(grand_label)
        grand_row_num = current_row_idx + 2
        row_to_sum_rows[grand_row_num] = all_month_row_nums
    sheet9 = wb.create_sheet('Summary')
    headers = ['Period', 'IDR', 'USD', 'Qty']
    data_2d = [[label] + [''] * 3 for label in period_labels]
    table_name = 'Summary'
    col_formats = {1: '@', 2: config.idr_format_2, 3: config.usd_format_2, 4: '#,##0'}
    row_formulas = {}
    for r_idx in range(len(period_labels)):
        pc = period_columns[r_idx] if r_idx < len(period_columns) else None
        r = r_idx + 2
        row_formulas[r_idx] = {}
        if pc and pc['type'] == 'month':
            per = pc['periods'][0]
            y, m = map(int, per.split('-'))
            date_str = f'DATE({y},{m},1)'
            row_formulas[r_idx][2] = f'=SUMIF(Sales_Data[Periode], {date_str}, Sales_Data[Total (IDR)])'
            row_formulas[r_idx][3] = f'=SUMIF(Sales_Data[Periode], {date_str}, Sales_Data[Total (USD)])'
            row_formulas[r_idx][4] = f'=SUMIF(Sales_Data[Periode], {date_str}, Sales_Data[Jumlah])'
        else:
            sum_rows = row_to_sum_rows.get(r, [])
            if sum_rows:
                sum_str_b = ','.join(f'B{sr}' for sr in sum_rows)
                row_formulas[r_idx][2] = f'=SUM({sum_str_b})'
                sum_str_c = ','.join(f'C{sr}' for sr in sum_rows)
                row_formulas[r_idx][3] = f'=SUM({sum_str_c})'
                sum_str_d = ','.join(f'D{sr}' for sr in sum_rows)
                row_formulas[r_idx][4] = f'=SUM({sum_str_d})'
    generate_table(sheet9, headers, data_2d, table_name, col_formats, row_formulas)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    buffer = list(output.getvalue())
    if blank_cells:
        msg = '<p>Processing complete. Warning: Please check these empty cells in the output:</p><ul>' + ''.join(f'<li>{c}</li>' for c in blank_cells) + '</ul>'
        message_type = 'warning'
    else:
        msg = 'Processing complete. All data processed successfully.'
        message_type = 'success'
    message_div.innerHTML = msg
    return {'buffer': buffer, 'message': msg, 'type': message_type}

window.process_files = ffi.create_proxy(process_files)